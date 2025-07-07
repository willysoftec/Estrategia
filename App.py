import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import plotly.express as px
from io import BytesIO

st.set_page_config(page_title="DJ - Contabilidad", layout="wide")
st.title("Sistema Contable DJ – Réplica web")

@st.cache_data
def load_workbook_obj(file):
    return load_workbook(file, keep_vba=False, data_only=True)

def sheet_to_df(ws):
    data = list(ws.values)
    return pd.DataFrame(data[1:], columns=data[0])

uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx", "xlsm"])

if uploaded_file:
    wb = load_workbook_obj(uploaded_file)
    sheet_names = [s for s in wb.sheetnames if s.upper() != "CONFIG"]

    tabs = st.tabs(sheet_names)

    for tab, name in zip(tabs, sheet_names):
        with tab:
            df = sheet_to_df(wb[name])
            st.subheader(name)
            st.dataframe(df, use_container_width=True)

            if name.upper() == "INGRESOS Y GASTOS" and {"ENERO","GASTOS"}.issubset(df.columns):
                st.plotly_chart(px.bar(df, x="ENERO", y="GASTOS"), use_container_width=True)

            if name.upper() == "TRABAJADORES":
                try:
                    st.success(f"Total anual en salarios: {df.iloc[-1, -1]:,.2f}")
                except IndexError:
                    st.warning("No se encontró la celda de total anual.")

    # --- Exportar copia procesada ---
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for name in sheet_names:
                sheet_to_df(wb[name]).to_excel(writer, sheet_name=name, index=False)
        st.download_button("Descargar Excel procesado", data=buffer.getvalue(),
                           file_name="DJ_Procesado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("👈 Carga tu archivo para comenzar.")
